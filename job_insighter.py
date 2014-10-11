#!/usr/bin/python
#-*-coding:utf-8-*-
#
# by Jing Wang, Sep 2014
#
import sys
import urllib2
from StringIO import StringIO
import gzip
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

try:
    import simplejson as json
except ImportError:
    import json

baseLink = "https://data.usajobs.gov"
allLink = "https://data.usajobs.gov/api/jobs?GradeLow=01&GradeHigh=15"
oneLink = "https://data.usajobs.gov/api/jobs?GradeLow=01&GradeHigh=15&Page=%d"
roleLink = "https://data.usajobs.gov/api/jobs?Title=%s"
locationLink = "https://data.usajobs.gov/api/jobs?LocationName=%s"

def find_key(dic, val):
    """return the key of dictionary dic given the value"""
    return [k for k, v in dic.iteritems() if v == val]

def load_page(url , proxy=False):
    print url
    if proxy == True:
        proxy_address = ""
        proxy_port = ""
        if len(proxy_address):
            us_proxy = "http://"+proxy_address+":"+proxy_port
            print 'Using proxy: ' + us_proxy
            proxy_handler = urllib2.ProxyHandler({'http':us_proxy})
            opener = urllib2.build_opener(proxy_handler)
            urllib2.install_opener(opener)

    req = urllib2.Request(url)
    req.add_header('Accept-encoding', 'gzip')
    response = urllib2.urlopen(req, timeout=10)

    if response.info().get('Content-Encoding') == 'gzip':
        buf = StringIO(response.read())
        f = gzip.GzipFile(fileobj=buf)
        link = f.read()
    else:
        link = response.read()

    response.close()
    return link

import pdb
if len(sys.argv) < 2:
    print "\nUsage: ./job_insighter.py <option> <keyword> <output>\n"
    print "-A: All jobs option"
    print "-J: Unfilled Jobs"
    print "-L: Unfilled Locations"
    print "-O: Unfilled Agencies"
else:
    try:
	opt = str(sys.argv[1])
	param = str(sys.argv[2])
    except:
	pass
    if cmp(opt, '-A') == 0:
	page = load_page(allLink)
	data = json.loads(page)
	wb = Workbook()

	dest_filename = r'output.xlsx'

	ws0 = wb.active

	ws0.title = "All jobs"

	ws0.header_footer.center_header.text = 'All Jobs'
	ws0.header_footer.center_header.font_size = 14
	ws0.header_footer.center_header.font_name = "Tahoma,Bold"
	ws0.header_footer.center_header.font_color = "CC3366"

	row_number = 0
	for idx in range(1, int(data['Pages'])):
            page = load_page(oneLink%(idx))
            data = json.loads(page)

	    for row in data['JobData']:
                i = 1
                size_of_column = len(row)
                for item in row:
                    val = str(row[item].encode('utf-8'))
                    if row_number == 0:
                        ws0.cell(row = 1, column = i).value = item
                        i += 1
                        if size_of_column + 1 == i:
                            row_number = 1
                    else:
                        ws0.cell(row = row_number, column = i).value = val
                        i += 1
		try:
		    wb.save(filename = dest_filename)
		except:
		    pass
		row_number += 1
    if cmp(opt, '-J') == 0:
	dest_filename = r'output.xlsx'
	wb = load_workbook(filename = dest_filename)
	ws0 = wb.active
	ws1 = wb.create_sheet()
	ws1.title = "Unfilled Jobs"

	items = {}
	row_idx = 2
	job_title_col_idx = 8

	while ws0.cell(row = row_idx, column = job_title_col_idx).value is not None:
	    job_title = ws0.cell(row = row_idx, column = job_title_col_idx).value
	    job_title = str(job_title.encode('utf-8'))
	    row_idx += 1

	    url = roleLink%(job_title)
	    if ' ' in url:
		url = url.replace(" ", '%20')

	    try:
		page = load_page(url)
		data = json.loads(page)
	    except:
		pass

	    items[job_title] = int(data['TotalJobs'])

	vals = items.values()
	top_freq = max(vals)

	ws1.cell(row = 1, column = 1).value = r'JobTitle'
	ws1.cell(row = 1, column = 2).value = r'Frequency'
	row_number = 2

	for x in find_key(items, top_freq):
	    ws1.cell(row = row_number, column = 1).value = x
	    ws1.cell(row = row_number, column = 2).value = items[x]

	    try:
	        wb.save(filename = dest_filename)
	    except:
	        pass
	    row_number += 1

    if cmp(opt, '-L') == 0:
	dest_filename = r'output.xlsx'
	wb = load_workbook(filename = dest_filename)
	ws0 = wb.active
	ws1 = wb.create_sheet()
	ws1.title = "Unfilled Locations"

	items = {}
	row_idx = 2
	loc_col_idx = 9

	while ws0.cell(row = row_idx, column = loc_col_idx).value is not None:
	    locations = ws0.cell(row = row_idx, column = loc_col_idx).value
	    locations = str(locations.encode('utf-8'))
	    row_idx += 1

	    url = locationLink%(locations)
	    if len(url) > 1000:
		continue
	    if ' ' in url:
		url = url.replace(" ", '%20')

	    try:
		page = load_page(url)
		data = json.loads(page)
	    except:
		continue

	    items[locations] = int(data['TotalJobs'])

	vals = items.values()
	top_freq = max(vals)

	ws1.cell(row = 1, column = 1).value = r'Locations'
	ws1.cell(row = 1, column = 2).value = r'Frequency'
	row_number = 2

	for x in find_key(items, top_freq):
	    ws1.cell(row = row_number, column = 1).value = x
	    ws1.cell(row = row_number, column = 2).value = items[x]

	    try:
	        wb.save(filename = dest_filename)
	    except:
	        pass
	    row_number += 1

    if cmp(opt, '-O') == 0:
	dest_filename = r'output.xlsx'
	wb = load_workbook(filename = dest_filename)
	ws0 = wb.active
	ws1 = wb.create_sheet()
	ws1.title = "Unfilled Agencies"

	items = {}
	agencies = []
	row_idx = 2
	agency_col_idx = 10

	while ws0.cell(row = row_idx, column = agency_col_idx).value is not None:
	    agency = ws0.cell(row = row_idx, column = agency_col_idx).value
	    agency = str(agency.encode('utf-8'))
	    row_idx += 1

	    agencies.append(agency)

	for i in agencies:
	    items[i] = agencies.count(i)

	vals = items.values()
	top_freq = max(vals)

	ws1.cell(row = 1, column = 1).value = r'AgencySubElement'
	ws1.cell(row = 1, column = 2).value = r'Frequency'
	row_number = 2

	for x in find_key(items, top_freq):
	    ws1.cell(row = row_number, column = 1).value = x
	    ws1.cell(row = row_number, column = 2).value = items[x]

	    try:
	        wb.save(filename = dest_filename)
	    except:
	        pass
	    row_number += 1